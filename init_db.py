# -*- coding: utf-8 -*-
"""初始化数据库并导入历史数据（一次性执行，自动去重）"""

import sqlite3
import os
import glob
from datetime import datetime

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
DB_PATH = os.path.join(BASE_DIR, "ai_news.db")

PRIORITY_MAP = {
    "应用落地与工具": 0,
    "Agent智能体": 0,
    "模型技术": 1,
    "方法论与研究": 1,
    "行业生态与政策": 2,
}

COL_MAPPING = {
    "发布日期": "date",
    "发布平台": "platform",
    "消息分类": "category",
    "消息概要": "summary",
    "中文翻译原文": "cn_text",
    "原文网址": "url",
    "英文原文内容": "en_text",
}


def init_database():
    conn = sqlite3.connect(DB_PATH)
    conn.execute("""
        CREATE TABLE IF NOT EXISTS news (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            date TEXT NOT NULL,
            platform TEXT NOT NULL,
            category TEXT NOT NULL,
            priority INTEGER NOT NULL DEFAULT 2,
            summary TEXT,
            cn_text TEXT,
            url TEXT UNIQUE NOT NULL,
            en_text TEXT,
            likes INTEGER NOT NULL DEFAULT 0,
            created_at TEXT NOT NULL DEFAULT (datetime('now', 'localtime'))
        )
    """)
    conn.execute("CREATE INDEX IF NOT EXISTS idx_pd ON news(priority ASC, date DESC)")
    conn.execute("CREATE INDEX IF NOT EXISTS idx_date ON news(date DESC)")
    conn.execute("CREATE INDEX IF NOT EXISTS idx_cat ON news(category)")
    conn.commit()
    return conn


def import_excel(conn, xlsx_path):
    import openpyxl

    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    ws = wb.active

    # 读取表头
    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    col = {}
    for i, h in enumerate(headers):
        if h in COL_MAPPING:
            col[COL_MAPPING[h]] = i

    if "url" not in col:
        print("  [WARN] 未找到「原文网址」列，无法导入")
        wb.close()
        return 0, 0

    inserted = 0
    skipped = 0

    for row in ws.iter_rows(min_row=2, values_only=True):
        url = row[col["url"]]
        if not url or not str(url).strip():
            skipped += 1
            continue

        date_val = row[col["date"]]
        if isinstance(date_val, datetime):
            date_val = date_val.strftime("%Y-%m-%d")
        else:
            date_val = str(date_val) if date_val else ""

        category = str(row[col["category"]] or "")
        priority = PRIORITY_MAP.get(category, 2)

        try:
            cursor = conn.execute(
                "INSERT OR IGNORE INTO news "
                "(date, platform, category, priority, summary, cn_text, url, en_text) "
                "VALUES (?, ?, ?, ?, ?, ?, ?, ?)",
                (
                    date_val,
                    str(row[col["platform"]] or ""),
                    category,
                    priority,
                    str(row[col["summary"]] or ""),
                    str(row[col["cn_text"]] or ""),
                    str(url).strip(),
                    str(row[col["en_text"]] or ""),
                ),
            )
            if cursor.rowcount > 0:
                inserted += 1
            else:
                skipped += 1
        except Exception as e:
            print(f"  [WARN] {e}")
            skipped += 1

    wb.close()
    return inserted, skipped


if __name__ == "__main__":
    print("=" * 50)
    print("  每日AI前沿 - 数据库初始化")
    print("=" * 50)

    conn = init_database()

    # 查找最新的 Excel 文件
    xlsx_files = sorted(
        glob.glob(os.path.join(BASE_DIR, "AI前沿资讯_*.xlsx"))
    )

    if xlsx_files:
        latest = xlsx_files[-1]
        print(f"\n  读取文件: {os.path.basename(latest)}")
        inserted, skipped = import_excel(conn, latest)
        conn.commit()
        conn.close()
        print(f"  导入完成: +{inserted} 新增, {skipped} 去重跳过")
    else:
        conn.close()
        print("\n  未找到 Excel 历史文件，已创建空数据库")

    print(f"\n  数据库位置: {DB_PATH}")
    print("=" * 50)
