# -*- coding: utf-8 -*-
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "AI前沿资讯"

# ============ STYLES ============
header_font = Font(name="Microsoft YaHei", bold=True, size=12, color="FFFFFF")
header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

cell_font = Font(name="Microsoft YaHei", size=10)
cell_alignment = Alignment(vertical="top", wrap_text=True)
date_alignment = Alignment(horizontal="center", vertical="top")
category_alignment = Alignment(horizontal="center", vertical="top", wrap_text=True)
priority_alignment = Alignment(horizontal="center", vertical="top", wrap_text=True)

thin_border = Border(
    left=Side(style="thin", color="B0B0B0"),
    right=Side(style="thin", color="B0B0B0"),
    top=Side(style="thin", color="B0B0B0"),
    bottom=Side(style="thin", color="B0B0B0"),
)

alt_fill = PatternFill(start_color="E8F0FE", end_color="E8F0FE", fill_type="solid")

# ============ HEADERS ============
headers = ["发布日期", "发布平台", "消息分类", "消息概要", "中文翻译原文", "原文网址", "英文原文内容"]
for col_idx, header in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col_idx, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = header_alignment
    cell.border = thin_border

# ============ COLUMN WIDTHS ============
col_widths = [14, 22, 20, 50, 80, 60, 80]
for i, w in enumerate(col_widths, 1):
    ws.column_dimensions[get_column_letter(i)].width = w

# ============ DATA (pre-sorted by priority P0 -> P1 -> P2, then by date desc) ============
data = [
    # ====== P0: 应用落地与工具 / Agent智能体 ======
    {
        "date": "2026-05-08",
        "platform": "NVIDIA Developer Blog",
        "category": "Agent智能体",
        "summary": "NVIDIA Dynamo 发布多轮 Agent 推理与工具调用支持，新增流式工具分发（Streaming Tool Dispatch），工具调用可在解析完成后立即执行而非等待整轮结束。Claude Code 和 Codex 等框架已可直连 Dynamo 后端，SWE-Bench 验证通过。",
        "cn_text": (
            "NVIDIA 发布 Dynamo 推理框架的重大更新，新增多轮 Agent 推理与工具调用的完整支持。\n\n"
            "核心改进：\n"
            "- 流式工具分发（--enable-streaming-tool-dispatch）：工具调用在结构解析完成后立即通过 SSE 侧通道分发执行，无需等待整轮对话结束\n"
            "- Anthropic Preamble 剥离（--strip-anthropic-billing-header）：移除会话级计费头以恢复 KV Cache 复用，TTFT 降低约 5 倍（911ms → 169ms）\n"
            "- 推理+工具调用交错解析：修复了此前将推理和工具调用错误分组的问题，正确保留 reasoning → tool_call 的交错序列\n"
            "- 独立可复用解析器层：dynamo-protocols、dynamo-parsers、dynamo-tokenizers 作为独立 crate 发布\n\n"
            "框架兼容性：\n"
            "- Claude Code（Anthropic Messages API）：完全兼容，配置 ANTHROPIC_BASE_URL 即可\n"
            "- Codex（OpenAI Responses API）：支持请求压缩，需匹配正确的模型 Catalog Profile\n"
            "- OpenClaw（Anthropic Messages API）：共享 Anthropic 端点兼容性\n\n"
            "关键发现：模型 Catalog 元数据（基础指令、工具注册、推理参数等）对 Agent 行为的影响与 API Schema 合规性同等重要。"
        ),
        "url": "https://developer.nvidia.com/blog/streaming-tokens-and-tools-multi-turn-agentic-harness-support-in-nvidia-dynamo/",
        "en_text": (
            "NVIDIA announced multi-turn agentic harness support in Dynamo, enabling structured interactions where assistant turns interleave reasoning with tool calls. "
            "Key improvements: streaming tool dispatch (--enable-streaming-tool-dispatch) executes tools immediately after parsing; "
            "Anthropic preamble stripping restores KV cache reuse (~5x TTFT improvement); "
            "correct interleaved reasoning and tool-call parsing via PR #7358.\n\n"
            "Claude Code and Codex can connect directly to Dynamo backend. "
            "Model catalog metadata parity matters as much as API schema compliance for frameworks like Codex. "
            "Standalone parser crates (dynamo-protocols, dynamo-parsers, dynamo-tokenizers) are now available."
        ),
    },
    {
        "date": "2026-05-08",
        "platform": "DeepLearning.AI",
        "category": "Agent智能体",
        "summary": "DeepLearning.AI 联合 CopilotKit 发布免费课程「构建交互式 Agent 与生成式 UI」，教授开发者如何让 Agent 渲染图表、表单、白板等富交互组件。课程涵盖 AG-UI 协议、MCP Apps 集成和画布应用开发，8 节课+5 个代码示例。",
        "cn_text": (
            "DeepLearning.AI 与 CopilotKit 联合发布全新免费课程「Build Interactive Agents with Generative UI」。\n\n"
            "课程概要：\n"
            "- 时长：1 小时 25 分钟\n"
            "- 讲师：Atai Barkai（CopilotKit CEO，AG-UI 协议核心推动者）\n"
            "- 8 节视频课 + 5 个代码示例 + 1 个测验\n\n"
            "核心教学内容 —— 生成式 UI 三层架构：\n"
            "1. 受控式生成 UI（Controlled Gen UI）：定义自定义组件（饼图、航班卡片等），由 Agent 按需渲染\n"
            "2. 声明式生成 UI（Declarative Gen UI）：使用 A2UI 开放规范（与 Google 共同开发），Agent 从可复用组件目录中组装布局\n"
            "3. 开放式生成 UI（Open-Ended Gen UI）：连接 MCP Apps，Agent 从零创建完全开放的交互界面\n\n"
            "技术栈：CopilotKit（Agent 前端框架）、AG-UI 协议（开源）、LangChain/LangGraph（Agent 后端）、"
            "Google ADK Agent、A2UI、MCP Apps、React\n\n"
            "最终项目：构建画布应用，Agent 与用户在共享数据上协作工作。"
        ),
        "url": "https://www.deeplearning.ai/short-courses/build-interactive-agents-with-generative-ui/",
        "en_text": (
            "DeepLearning.AI and CopilotKit released a free course 'Build Interactive Agents with Generative UI' taught by Atai Barkai (CopilotKit CEO). "
            "Duration: 1h25m, 8 video lessons + 5 code examples + quiz.\n\n"
            "Covers the Generative UI Spectrum: Controlled Gen UI (custom components rendered on demand), "
            "Declarative Gen UI (A2UI open spec co-developed with Google), and Open-Ended Gen UI (MCP Apps). "
            "Tech stack: CopilotKit, AG-UI protocol, LangChain/LangGraph, Google ADK Agent, MCP Apps, React.\n\n"
            "Capstone: build a canvas application where agent and user work on shared data side by side. "
            "All skills built on the open-source AG-UI protocol for portability across the agentic ecosystem."
        ),
    },
    {
        "date": "2026-05-08",
        "platform": "NVIDIA Developer Blog",
        "category": "应用落地与工具",
        "summary": "NVIDIA 发布模型量化工具 Model Optimizer 的教程，支持训练后量化（Post-Training Quantization）降低 VRAM 占用并提升推理性能，可在 GeForce RTX 等消费级 GPU 上运行大型模型。",
        "cn_text": (
            "NVIDIA 发布关于使用 Model Optimizer 进行训练后量化（Post-Training Quantization）的详细教程。\n\n"
            "核心内容：\n"
            "- 面向消费级 GPU（如 GeForce RTX）的模型量化方案\n"
            "- 通过降低模型精度减少 VRAM 占用，使大模型可在显存有限的 GPU 上运行\n"
            "- 量化后保持推理精度同时提升吞吐量\n\n"
            "适用场景：开发者希望在本地硬件上部署 LLM 或视觉模型，"
            "但受限于消费级 GPU 的显存容量。"
        ),
        "url": "https://developer.nvidia.com/blog/model-quantization-post-training-quantization-using-nvidia-model-optimizer/",
        "en_text": (
            "NVIDIA published a guide on post-training quantization using NVIDIA Model Optimizer to reduce VRAM usage "
            "and improve inference performance on consumer GPUs like GeForce RTX. "
            "The tutorial covers quantization techniques that allow large models to run on hardware with limited memory."
        ),
    },
    {
        "date": "2026-05-08",
        "platform": "NVIDIA Developer Blog",
        "category": "应用落地与工具",
        "summary": "NVIDIA 探索通过语法约束解码（Grammar-Constrained Decoding）提升小语言模型生成有效 Bash 命令的能力。Bash 被描述为「暴露给 AI Agent 最灵活、最强大的接口之一」，此举可显著提升代码 Agent 的可靠性。",
        "cn_text": (
            "NVIDIA 发表技术文章，探索如何通过语法约束解码（Grammar-Constrained Decoding）提升小语言模型生成有效 Bash 命令的能力。\n\n"
            "背景：Bash 是 AI Agent 可用的最灵活、最强大的接口之一。Agent 通过 Bash 执行文件操作、系统命令和工具调用。"
            "但小模型在生成复杂 Bash 命令时容易产生语法错误。\n\n"
            "解决方案：\n"
            "- 利用语法约束解码技术，在推理时强制模型输出符合 Bash 语法的合法命令\n"
            "- 限制模型输出空间，消除无效 token 的生成可能\n"
            "- 使更小、更高效的模型也能可靠地执行 Bash 操作\n\n"
            "价值：降低对大型模型的依赖，使轻量级 Agent 在边缘设备或低资源环境中也能可靠运行。"
        ),
        "url": "https://developer.nvidia.com/blog/improving-bash-generation-in-small-language-models-with-grammar-constrained-decoding/",
        "en_text": (
            "NVIDIA published a technical article on improving Bash command generation in small language models using grammar-constrained decoding. "
            "Bash is described as 'one of the most flexible and powerful interfaces exposed to AI agents.' "
            "Grammar-constrained decoding restricts the model's output space during inference to ensure syntactically valid Bash commands, "
            "enabling smaller, more efficient models to reliably execute Bash operations."
        ),
    },
    {
        "date": "2026-05-08",
        "platform": "OpenAI Academy",
        "category": "Agent智能体",
        "summary": "OpenAI Academy 发布「Workspace Agents」企业管理工作坊，面向 ChatGPT Enterprise 管理员，讲解 Agent 在企业工作空间中的部署、管理与最佳实践。此外还推出了 ChatGPT for Excel 技能实验室和高校教师 AI 应用工作坊。",
        "cn_text": (
            "OpenAI Academy 发布多个最新活动和工作坊：\n\n"
            "1. Workspace Agents: ChatGPT Enterprise 管理员指南（5 月 19 日）\n"
            "   - 面向企业级管理员的 Agent 部署与管理指导\n"
            "   - 涵盖 Agent 在企业工作空间中的安全配置与最佳实践\n\n"
            "2. Skill Lab: Using ChatGPT for Excel（5 月 14 日）\n"
            "   - 30 分钟实操技能实验室\n"
            "   - 学习如何用 ChatGPT 增强 Excel 工作流\n\n"
            "3. ChatGPT for Higher Education Faculty（5 月 14 日）\n"
            "   - 面向高校教师的 ChatGPT 学术应用工作坊\n"
            "   - 探讨 AI 在教学和科研中的实际应用\n\n"
            "OpenAI Academy 还宣布认证项目（Certifications）正在进行中，"
            "将提供从 Prompt Engineering 基础到 AI 工作全流程的不同级别认证。"
        ),
        "url": "https://academy.openai.com/en/public/events/workspace-agents-guidance-for-chatgpt-enterprise-admins-zs4dny9yey",
        "en_text": (
            "OpenAI Academy announced multiple upcoming events:\n"
            "- 'Workspace Agents: Guidance for ChatGPT Enterprise Admins' (May 19) — enterprise admin guidance on agent deployment\n"
            "- 'Skill Lab: Using ChatGPT for Excel' (May 14) — 30-min hands-on lab on enhancing Excel workflows\n"
            "- 'ChatGPT for Higher Education Faculty' (May 14) — workshop on academic AI applications\n\n"
            "OpenAI Certifications program is in pilot phase, planning certifications at different AI fluency levels."
        ),
    },
    {
        "date": "2026-05-05",
        "platform": "Anthropic Academy",
        "category": "Agent智能体",
        "summary": "Anthropic Academy 持续扩展课程体系，新增 Agent Skills（构建可复用 Markdown 指令）和 Subagents（上下文管理与任务委托）两大 Claude Code 课程，以及 Introduction to Claude Cowork（任务循环、插件与文件工作流）课程。",
        "cn_text": (
            "Anthropic Academy 持续扩展其课程体系，目前共 17 门课程。新增的课程亮点：\n\n"
            "1. Introduction to Agent Skills\n"
            "   - 学习构建、配置和共享 Skills —— 可复用的 Markdown 指令\n"
            "   - Claude Code 自动应用这些指令来增强工作能力\n\n"
            "2. Introduction to Subagents\n"
            "   - 学习使用和创建子 Agent 来管理上下文、委托任务和构建专业工作流\n"
            "   - 在 Claude Code 中实现多 Agent 协作\n\n"
            "3. Introduction to Claude Cowork\n"
            "   - 涵盖 Cowork 任务循环、插件与技能系统\n"
            "   - 文件与研究工作流\n"
            "   - 负责任地引导多步工作\n\n"
            "此外还有进阶课程：MCP 高级主题（采样、通知、文件系统访问、传输机制）、"
            "Claude with Amazon Bedrock、Claude with Google Vertex AI 等。\n\n"
            "所有课程均免费开放，无需 Anthropic 账号。"
        ),
        "url": "https://anthropic.skilljar.com",
        "en_text": (
            "Anthropic Academy expanded to 17 courses covering the full Claude ecosystem. "
            "Notable newer courses include: Introduction to Agent Skills (building reusable markdown instructions for Claude Code), "
            "Introduction to Subagents (context management and task delegation), and Introduction to Claude Cowork "
            "(task loop, plugins, file/research workflows). "
            "Advanced topics: MCP Advanced Topics, Claude with Amazon Bedrock, Claude with Google Vertex AI. "
            "All courses are free and require only a Skilljar account."
        ),
    },
    {
        "date": "2026-05-05",
        "platform": "NVIDIA Developer Blog",
        "category": "Agent智能体",
        "summary": "NVIDIA 发布「从云端到汽车」的车载 AI Agent 构建指南，描述从规则式界面到 Agent 式多模态 AI 系统的转变。NVIDIA 提出的方案支持 Agent 进行推理、规划和自主行动，面向智能座舱场景。",
        "cn_text": (
            "NVIDIA 发布了关于构建车载 AI Agent 的技术指南，覆盖从云端到汽车的完整方案。\n\n"
            "核心转变：\n"
            "- 从传统的规则式界面转向 Agent 式多模态 AI 系统\n"
            "- 座舱 AI 需要具备推理、规划和自主行动能力\n\n"
            "技术方案要点：\n"
            "- 多模态输入处理（语音、视觉、触控等）\n"
            "- 云端-车端协同推理架构\n"
            "- 安全关键场景下的 Agent 行为约束\n\n"
            "应用场景：智能座舱交互、驾驶辅助信息呈现、车机个性化服务。"
        ),
        "url": "https://developer.nvidia.com/blog/how-to-build-in-vehicle-ai-agents-with-nvidia-from-cloud-to-car/",
        "en_text": (
            "NVIDIA published a guide on building in-vehicle AI agents from cloud to car. "
            "The article describes the shift from rule-based interfaces to agentic, multimodal AI systems in automotive cockpits. "
            "NVIDIA's approach covers multimodal input processing, cloud-car collaborative inference architecture, "
            "and agent behavior constraints for safety-critical scenarios."
        ),
    },

    # ====== P1: 模型技术 / 方法论与研究 ======
    {
        "date": "2026-05-07",
        "platform": "Google Grow with Google",
        "category": "模型技术",
        "summary": "Google 发布 AI Professional Certificate（AI 专业认证），包含 7 门课程的完整体系，涵盖 AI 基础、头脑风暴与研究、数据分析、内容创作、应用构建（Vibe Coding）等模块。注册即获 90 天 Google AI Pro 免费使用权。",
        "cn_text": (
            "Google 发布全新 AI Professional Certificate（AI 专业认证），面向职场人士的系统化 AI 技能培养项目。\n\n"
            "课程体系（7 门课程）：\n"
            "1. AI Fundamentals（AI 基础）\n"
            "2. AI for Brainstorming and Planning（头脑风暴与规划）\n"
            "3. AI for Research and Insights（研究与洞察）\n"
            "4. AI for Writing and Communicating（写作与沟通）\n"
            "5. AI for Content Creation（内容创作）\n"
            "6. AI for Data Analysis（数据分析）\n"
            "7. AI for App Building（应用构建 —— Vibe Coding，无需编程）\n\n"
            "特色：\n"
            "- 20+ 个实操练习，建立 AI 实战能力\n"
            "- 由美国大型雇主联盟共同认证\n"
            "- 注册即可获得 90 天 Google AI Pro 免费使用权\n"
            "- 帮助学员构建求职作品集\n\n"
            "行业数据：AI 相关岗位过去两年增长 108%，掌握 AI 技能的员工工资溢价达 56%。"
        ),
        "url": "https://grow.google/ai-professional/",
        "en_text": (
            "Google launched the AI Professional Certificate, a comprehensive 7-course program covering AI Fundamentals, "
            "Brainstorming and Planning, Research and Insights, Writing and Communicating, Content Creation, "
            "Data Analysis, and App Building (vibe coding — no coding required). "
            "Features 20+ hands-on activities, validated by a coalition of major U.S. employers, "
            "and includes 90 days of free Google AI Pro. "
            "AI-related job postings increased 108% in the past 2 years; AI-skilled workers earn a 56% wage premium."
        ),
    },
    {
        "date": "2026-05-08",
        "platform": "Hugging Face Learn",
        "category": "方法论与研究",
        "summary": "Hugging Face 持续扩展开源 AI 课程体系，当前 12 门免费课程覆盖 LLM、Agent、上下文工程（Code Agent 上下文工程）、机器人学（LeRobot）、后训练（Post-Training）、扩散模型、3D ML 等前沿领域。上下文工程和 Agent 课程为近期新增重点。",
        "cn_text": (
            "Hugging Face Learn 平台持续扩展其免费开源 AI 课程体系，目前共 12 门课程：\n\n"
            "核心课程：\n"
            "- LLM Course（大语言模型）\n"
            "- Context Course（上下文工程 —— 针对 Code Agent 的新兴方向）\n"
            "- Agents Course（AI Agent 构建与部署）\n"
            "- a smol course（后训练 AI 模型的入门课程）\n\n"
            "应用领域课程：\n"
            "- Robotics Course（使用 LeRobot 构建机器人）\n"
            "- Computer Vision Course（计算机视觉）\n"
            "- Audio Course（语音/音频处理）\n"
            "- Diffusion Course（扩散模型与 Diffusers）\n"
            "- ML for 3D Course（3D 机器学习）\n"
            "- ML for Games Course（AI 在游戏开发中的应用）\n\n"
            "进阶课程：\n"
            "- Deep RL Course（深度强化学习）\n"
            "- Open-Source AI Cookbook（开源 AI 食谱）\n\n"
            "亮点：Context Course 和 Agents Course 是当前最热门的新增方向，"
            "反映了 AI 行业从单纯的大模型开发向 Agent 工程和上下文优化的趋势转变。"
            "所有课程完全免费，基于 Hugging Face 生态系统。"
        ),
        "url": "https://huggingface.co/learn",
        "en_text": (
            "Hugging Face Learn platform offers 12 free open-source courses covering LLMs, Context Engineering (for Code Agents), "
            "AI Agents, Post-Training, Robotics (LeRobot), Computer Vision, Audio, Diffusion Models, "
            "3D ML, Games, Deep RL, and an AI Cookbook. "
            "The Context Course and Agents Course represent the newest additions reflecting the industry shift "
            "toward agent engineering and context optimization. All courses are free and built on the Hugging Face ecosystem."
        ),
    },
    {
        "date": "2026-05-08",
        "platform": "DeepLearning.AI",
        "category": "方法论与研究",
        "summary": "Andrew Ng 在 The Batch 第 352 期发表专栏文章「不会有 AI 就业末日」，反驳 AI 导致大规模失业的观点。指出 AI 受影响最大的软件工程行业招聘依然强劲，美国失业率保持 4.3%，AI 创造的就业远超被替代的岗位。",
        "cn_text": (
            "Andrew Ng 在 DeepLearning.AI 的 The Batch 第 352 期发表署名文章，强烈反驳「AI 就业末日」论。\n\n"
            "核心论点：\n"
            "- 软件工程是受 AI 工具影响最大的行业，但招聘依然强劲\n"
            "- 美国失业率保持 4.3% 的健康水平\n"
            "- AI 净创造的就业远超被替代的岗位，与历次技术革命一致\n\n"
            "「末日叙事」持续存在的原因：\n"
            "1. 前沿 AI 实验室有动机夸大 AI 能力（锚定 10 万美元年薪来定价 SaaS 产品）\n"
            "2. 企业有动机将裁员归因于 AI（掩盖疫情期间低利率刺激下的过度招聘）\n\n"
            "Andrew Ng 预测：\n"
            "「将迎来 AI 就业嘉年华（AI Jobapalooza）！AI 将创造大量优质的 AI 工程岗位，"
            "我对整体就业市场的未来持乐观态度。」\n\n"
            "同期新闻还包括：ByteDance Seedance 2.0 视频生成模型上线 CapCut、"
            "NVIDIA AI 引导芯片设计、Gallup 调查显示 50% 美国工作者在工作中使用 AI、"
            "解决机器人 VLA 模型灾难性遗忘问题的研究。"
        ),
        "url": "https://www.deeplearning.ai/the-batch/issue-352",
        "en_text": (
            "Andrew Ng published 'There Will Be No AI Jobpocalypse' in The Batch Issue 352. "
            "Key arguments: software engineering (most affected by AI) still shows strong hiring; U.S. unemployment at 4.3%; "
            "net AI job creation vastly exceeds job destruction.\n\n"
            "Ng explains the jobpocalypse narrative persists because: (1) frontier AI labs incentivize making AI sound more powerful, "
            "and (2) businesses attribute layoffs to AI rather than admitting pandemic overhiring.\n\n"
            "Predicts an 'AI Jobapalooza' with abundant good AI engineering jobs. "
            "Same issue covers: ByteDance Seedance 2.0 on CapCut, Nvidia's AI-guided chip design, "
            "Gallup poll (50% of U.S. workers use AI at work), and research on mitigating catastrophic forgetting in robot VLA models."
        ),
    },
    {
        "date": "2026-05-07",
        "platform": "NVIDIA Developer Blog",
        "category": "模型技术",
        "summary": "NVIDIA 发布 GB200 NVL72 系统性能优化指南，利用 Slurm Block Scheduling 实现峰值系统效率。GB200 NVL72 引入了跨整个机架的 NVLink 一致性，代表 GPU 集群架构的根本性转变。",
        "cn_text": (
            "NVIDIA 发布了关于在 GB200 NVL72 系统上使用 Slurm Block Scheduling 实现峰值性能的技术指南。\n\n"
            "技术背景：\n"
            "- GB200 NVL72 引入了跨整个机架的 NVLink 一致性\n"
            "- 这是 GPU 集群架构的根本性转变\n"
            "- 需要新的调度策略来最大化系统利用率\n\n"
            "优化方案：\n"
            "- 利用 Slurm Block Scheduling 优化 GPU 资源分配\n"
            "- 实现峰值系统和负载效率\n"
            "- 面向大规模分布式训练场景\n\n"
            "适用场景：需要在 GB200 NVL72 机架级系统上进行大规模 AI 训练的团队。"
        ),
        "url": "https://developer.nvidia.com/blog/achieving-peak-system-and-workload-efficiency-on-nvidia-gb200-nvl72-with-slurm-block-scheduling/",
        "en_text": (
            "NVIDIA published a guide on achieving peak system and workload efficiency on NVIDIA GB200 NVL72 using Slurm Block Scheduling. "
            "The GB200 NVL72 introduces NVLink coherence across an entire rack — a fundamental shift in GPU cluster architecture "
            "requiring new scheduling strategies to maximize system utilization."
        ),
    },
    {
        "date": "2026-05-07",
        "platform": "NVIDIA Developer Blog",
        "category": "模型技术",
        "summary": "NVIDIA 发布 NCCL Inspector 与 Prometheus 集成方案，支持分布式深度学习训练中 GPU 间通信的实时性能监控和更快的调试，面向大规模模型训练的性能调优需求。",
        "cn_text": (
            "NVIDIA 发布了 NCCL Inspector 与 Prometheus 集成的技术方案。\n\n"
            "核心功能：\n"
            "- 实时监控 GPU 间通信性能（NCCL 是 NVIDIA 集体通信库）\n"
            "- 通过 Prometheus 进行指标采集和可视化\n"
            "- 加速分布式训练中的通信瓶颈调试\n\n"
            "价值：在大规模分布式深度学习训练中，GPU 间通信往往是性能瓶颈。"
            "该工具使开发者能够实时观察通信延迟、带宽利用率和拥塞情况，"
            "快速定位和解决训练效率问题。"
        ),
        "url": "https://developer.nvidia.com/blog/real-time-performance-monitoring-and-faster-debugging-with-nccl-inspector-and-prometheus/",
        "en_text": (
            "NVIDIA introduced NCCL Inspector integration with Prometheus for real-time monitoring and debugging "
            "of GPU-to-GPU communication in distributed deep learning training. "
            "The tool enables developers to observe communication latency, bandwidth utilization, and congestion in real-time, "
            "accelerating performance bottleneck identification in large-scale distributed training."
        ),
    },
    {
        "date": "2026-05-05",
        "platform": "NVIDIA Developer Blog",
        "category": "方法论与研究",
        "summary": "NVIDIA 发文探讨 Agent 系统复杂度管理的「极致协同设计」（Extreme Co-Design）方法论。指出 Agent AI 与第一波生成式 AI 的本质区别 —— Agent 不只是响应，而是自主行动。需要在硬件、软件和系统层面进行协同设计。",
        "cn_text": (
            "NVIDIA 发布关于应对 Agent 系统复杂度上升的「极致协同设计」（Extreme Co-Design）方法论。\n\n"
            "核心观点：\n"
            "- Agent AI 与第一波生成式 AI 有本质区别：Agent 不只是响应，而是自主行动\n"
            "- Agent 系统的复杂度远超传统 LLM 应用\n"
            "- 需要在硬件、软件和系统架构层面进行协同设计\n\n"
            "方法论要点：\n"
            "- 硬件层：GPU、网络、存储针对 Agent 工作负载优化\n"
            "- 软件层：推理框架、调度器、工具调用协议的统一设计\n"
            "- 系统层：端到端的性能优化和资源管理\n\n"
            "这代表了 NVIDIA 对 AI Agent 基础设施的战略思考，"
            "表明行业正从「把模型做得更大」转向「让系统更好地协同工作」。"
        ),
        "url": "https://developer.nvidia.com/blog/building-for-the-rising-complexity-of-agentic-systems-with-extreme-co-design/",
        "en_text": (
            "NVIDIA discussed 'extreme co-design' for handling the rising complexity of agentic AI systems. "
            "Key insight: agentic AI differs from first-wave generative AI — agents don't just respond, they act autonomously. "
            "This requires co-design across hardware (GPU, networking, storage), software (inference frameworks, schedulers, tool-calling protocols), "
            "and system architecture (end-to-end optimization and resource management). "
            "The methodology signals the industry's shift from 'making models bigger' to 'making systems work better together'."
        ),
    },

    # ====== P2: 行业生态与政策 ======
    {
        "date": "2026-05-08",
        "platform": "IBM SkillsBuild",
        "category": "行业生态与政策",
        "summary": "IBM SkillsBuild 举办多场 AI 主题虚拟活动，包括 5 月 20 日「启动数据与 AI 之旅」、5 月 25 日「AI 基础：语言与视觉」及「Prompt Craft 与 AI 交互技能」工作坊，面向成年学习者和大学生。",
        "cn_text": (
            "IBM SkillsBuild 举办多场 AI 相关虚拟活动和培训：\n\n"
            "1. Kickstart your Data and AI journeys（5 月 20 日）\n"
            "   - 面向成年学习者和大学生的多日虚拟活动\n"
            "   - 覆盖数据分析和 AI 学习路径\n\n"
            "2. AI Fundamentals: Language and Vision in AI（5 月 25 日，APAC）\n"
            "   - 面向亚太地区的 AI 基础培训\n"
            "   - 涵盖 AI 语言理解和计算机视觉\n\n"
            "3. Prompt Craft and AI Interaction Track（5 月 25 日）\n"
            "   - 专注于 Prompt Engineering 和 AI 交互技能\n"
            "   - 面向成年学习者和大学生\n\n"
            "IBM 持续推进其「3 年内培训 200 万人掌握 AI」的承诺，"
            "并扩展至「2030 年前为全球 3000 万人提供技能培训」的宏大目标。"
        ),
        "url": "https://skillsbuild.org/events",
        "en_text": (
            "IBM SkillsBuild is hosting multiple AI-themed virtual events:\n"
            "- 'Kickstart your Data and AI journeys' (May 20) — multi-day event for adult learners and college students\n"
            "- 'AI Fundamentals: Language and Vision in AI' (May 25, APAC) — AI fundamentals for Asia-Pacific region\n"
            "- 'Prompt Craft and AI Interaction Track' (May 25) — prompt engineering and AI interaction skills\n\n"
            "IBM continues its commitment to training 2 million people in AI within 3 years "
            "and skilling 30 million people globally by 2030."
        ),
    },
    {
        "date": "2026-05-08",
        "platform": "DeepLearning.AI",
        "category": "行业生态与政策",
        "summary": "Gallup 最新民调显示 50% 美国工作者去年在工作中使用过 AI，日使用率从 2023 年的 4% 跃升至 13%。65% 的 AI 使用者称 AI 提升了生产力，但 31% 表示 AI 改变了他们的工作方式。",
        "cn_text": (
            "Gallup 最新民调（2026 年 2 月 4-19 日，23,700 名美国员工）揭示了 AI 在职场中的实际应用情况。\n\n"
            "核心数据：\n"
            "- 50% 的美国工作者去年在工作中至少使用过几次 AI\n"
            "- 日使用率从 2023 年的 4% 跃升至 13%\n"
            "- 每周使用几次的比例从 11% 提升至 28%\n"
            "- 2/5 的工作者表示雇主引入了 AI 工具\n"
            "- 25% 的公司拥有清晰的 AI 战略\n"
            "- 65% 的 AI 使用者称 AI 提升了生产力\n"
            "- 31% 表示 AI 改变了他们的工作方式\n\n"
            "矛盾信号：\n"
            "- Apollo 首席经济学家：「AI 无处不在，除了宏观经济数据中」\n"
            "- Stanford 研究：AI 影响岗位的就业率在下降\n"
            "- Brookings 研究：投资 AI 的公司反而招聘了更多员工\n\n"
            "结论：工作者用 AI 来辅助工作，而非替代工作。"
        ),
        "url": "https://www.deeplearning.ai/the-batch/issue-352",
        "en_text": (
            "Gallup poll of 23,700 U.S. employees (Feb 4-19, 2026) shows: 50% of U.S. workers used AI at work at least a few times in the past year; "
            "daily usage rose from 4% (2023) to 13%; weekly usage from 11% to 28%; "
            "65% of AI users report productivity improvements; 31% say AI changed how they work.\n\n"
            "Conflicting macro signals: Apollo chief economist notes 'AI is everywhere except in the incoming macroeconomic data'; "
            "Stanford research shows declining employment for AI-affected roles; "
            "Brookings study finds companies investing in AI hired more workers."
        ),
    },
    {
        "date": "2026-05-08",
        "platform": "DeepLearning.AI",
        "category": "行业生态与政策",
        "summary": "ByteDance 将 Seedance 2.0 多模态视频生成模型上线 CapCut（7.36 亿月活），支持文本/图片/音频/视频输入，生成 4-15 秒同步视频+音频。Arena AI 排行榜文生视频和图生视频双榜第一。好莱坞六大制片厂要求停止使用版权素材训练。",
        "cn_text": (
            "ByteDance 将 Seedance 2.0 多模态视频生成模型上线 CapCut 全球版（7.36 亿月活跃用户，仅次于 ChatGPT 的第二大消费级 AI 产品）。\n\n"
            "技术规格：\n"
            "- 输入：文本、图片（最多 9 张）、音频（最多 3 段）、视频（最多 3 段）\n"
            "- 输出：同步视频+音频，4-15 秒，480 或 720px\n"
            "- 6 种画面比例（21:9 到 9:16）\n"
            "- 功能：多语言对口型、环境音、音乐、多机位剪辑、隐形水印、人脸/版权屏蔽\n"
            "- 定价：$0.24-$0.30/秒\n\n"
            "排行榜表现：\n"
            "- Arena AI：文生视频和图生视频双榜第一（1,460/1,454 Elo）\n"
            "- Artificial Analysis：图生视频+同步音频排名第一\n\n"
            "争议：生成了一段模仿 Tom Cruise 和 Brad Pitt 的视频，"
            "引发好莱坞六大制片厂抗议。行业背景：OpenAI 关停 Sora（DAU 从 100 万降至 50 万以下）。"
        ),
        "url": "https://www.deeplearning.ai/the-batch/issue-352",
        "en_text": (
            "ByteDance added Seedance 2.0 to CapCut (736M MAU — 2nd largest consumer AI product after ChatGPT). "
            "Inputs: text, images (up to 9), audio (up to 3 clips), video (up to 3 clips). "
            "Outputs: synchronized video+audio, 4-15 seconds, 480/720px, 6 aspect ratios. "
            "Features: lip-synced dialogue, ambient sound, music, multi-camera cuts, invisible watermarking.\n\n"
            "Arena AI: #1 in text-to-video and image-to-video (1,460/1,454 Elo). "
            "Controversy: generated clip featuring Tom Cruise and Brad Pitt likenesses prompted demands from six Hollywood studios.\n\n"
            "Industry context: OpenAI shutting down Sora (DAUs fell from ~1M to under 500K)."
        ),
    },
]

# ============ WRITE DATA ============
for row_idx, item in enumerate(data, 2):
    values = [
        item["date"],
        item["platform"],
        item["category"],
        item["summary"],
        item["cn_text"],
        item["url"],
        item["en_text"],
    ]
    for col_idx, value in enumerate(values, 1):
        cell = ws.cell(row=row_idx, column=col_idx, value=value)
        cell.font = cell_font
        cell.border = thin_border
        if col_idx == 1:
            cell.alignment = date_alignment
        elif col_idx == 3:
            cell.alignment = category_alignment
        else:
            cell.alignment = cell_alignment
        # Alternate row coloring
        if row_idx % 2 == 0:
            cell.fill = alt_fill

# ============ FREEZE TOP ROW ============
ws.freeze_panes = "A2"

# ============ AUTO FILTER ============
ws.auto_filter.ref = f"A1:G{len(data) + 1}"

# ============ ROW HEIGHT ============
ws.row_dimensions[1].height = 30
for r in range(2, len(data) + 2):
    ws.row_dimensions[r].height = 180

# ============ SAVE ============
output_path = r"E:\每日AI前沿\AI前沿资讯_2026-05-09.xlsx"
wb.save(output_path)
print(f"Excel file saved to: {output_path}")
print(f"Total news items: {len(data)}")
